# -*- coding: utf-8 -*-
"""Excel (.xlsx) rapor üretici v4.

TB skora dahil değil. Genel Skor = %50 BIM + %50 QC.
Sheet Personnel Metadata ayrı sheet. openpyxl gerekli.
"""
from __future__ import division, print_function

import os

from qc_utils import ensure_folder, get_logger, utext

logger = get_logger()


def _try_import_openpyxl():
    try:
        import openpyxl
        return openpyxl, True
    except ImportError:
        return None, False


def export_xlsx_report(result, xlsx_path):
    openpyxl, available = _try_import_openpyxl()
    if not available:
        return False, u"openpyxl yüklü değil"

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    try:
        ensure_folder(os.path.dirname(xlsx_path))
        wb = openpyxl.Workbook()

        hf = Font(bold=True, size=11, color="FFFFFF")
        hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        tborder = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

        def sty_hdr(ws, row, mc):
            for c in range(1, mc + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = hf; cell.fill = hfill; cell.border = tborder

        def auto_w(ws, mc, mw=50):
            for c in range(1, mc + 1):
                ml = max((len(utext(cell.value)) for row in ws.iter_rows(min_col=c, max_col=c) for cell in row if cell.value), default=8)
                ws.column_dimensions[get_column_letter(c)].width = min(ml + 3, mw)

        # === SHEET 1: Executive Summary ===
        ws1 = wb.active
        ws1.title = "Executive Summary"
        gsd = result.get('general_score_data', {})
        rows_data = [
            (u"Proje", result.get('project_name', '')),
            (u"Dosya", result.get('document_title', '')),
            (u"Tarih", result.get('run_time', '')),
            (u"Disiplin", result.get('discipline_name', '')),
            (u"", u""),
            (u"GENEL SKOR", u"{0}/100".format(result.get('total_score', 0))),
            (u"BIM Formu Skoru", u"{0}/100 (ağırlık: %{1})".format(gsd.get('bim_form_score', 0), int(gsd.get('bim_weight', 0.5)*100))),
            (u"Diğer QC Skoru", u"{0}/100 (ağırlık: %{1})".format(gsd.get('other_qc_score', 0), int(gsd.get('qc_weight', 0.5)*100))),
            (u"Kalite Sınıfı", result.get('quality_class', '')),
            (u"Teslim Kararı", result.get('delivery_decision', '')),
            (u"Kırmızı Bayrak", len(result.get('red_flags', []))),
            (u"", u""),
            (u"NOT", u"Title block verileri kalite skoruna dahil değildir"),
        ]
        for i, (l, v) in enumerate(rows_data, 1):
            ws1.cell(row=i, column=1, value=l).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=utext(v))

        r = len(rows_data) + 2
        headers = [u"Kategori", u"Ağırlık", u"Puan (0-5)", u"Katkı", u"Açıklama"]
        for c, h in enumerate(headers, 1):
            ws1.cell(row=r, column=c, value=h)
        sty_hdr(ws1, r, len(headers))
        for rd in result.get('rows', []):
            r += 1
            ws1.cell(row=r, column=1, value=rd['category'])
            ws1.cell(row=r, column=2, value=rd['weight'])
            ws1.cell(row=r, column=3, value=rd['score5'])
            ws1.cell(row=r, column=4, value=rd['weighted_score'])
            ws1.cell(row=r, column=5, value=rd['note'])
        auto_w(ws1, 5)
        ws1.freeze_panes = 'A2'

        # === SHEET 2: BIM Form Compliance ===
        ws2 = wb.create_sheet("BIM Form Compliance")
        bh = [u"ID", u"Açıklama", u"Tip", u"Durum", u"Skor", u"Ağırlık", u"Kanıt", u"Manuel?", u"Kullanıcı Onayı"]
        for c, h in enumerate(bh, 1):
            ws2.cell(row=1, column=c, value=h)
        sty_hdr(ws2, 1, len(bh))
        for i, item in enumerate(result.get('bim_form_results', []), 2):
            ws2.cell(row=i, column=1, value=item.get('id', ''))
            ws2.cell(row=i, column=2, value=item.get('description', ''))
            ws2.cell(row=i, column=3, value=item.get('check_type', ''))
            ws2.cell(row=i, column=4, value=item.get('status', ''))
            sv = item.get('score', -1)
            ws2.cell(row=i, column=5, value=sv if sv >= 0 else u"Bekliyor")
            ws2.cell(row=i, column=6, value=item.get('weight', 0))
            ws2.cell(row=i, column=7, value=item.get('evidence', ''))
            ws2.cell(row=i, column=8, value=u"Evet" if item.get('needs_manual') else u"Hayır")
            ua = item.get('user_approved')
            ws2.cell(row=i, column=9, value=u"Evet" if ua is True else (u"Hayır" if ua is False else u"—"))
            if item.get('needs_manual'):
                for col in range(1, len(bh) + 1):
                    ws2.cell(row=i, column=col).fill = yellow
        auto_w(ws2, len(bh))
        ws2.freeze_panes = 'A2'
        ws2.auto_filter.ref = ws2.dimensions

        # === SHEET 3: Sheet Personnel Metadata (SKOR DIŞI) ===
        ws3 = wb.create_sheet("Sheet Personnel")
        ph = [u"Sheet Number", u"Sheet Name", u"Drawn By", u"Designed By",
              u"Checked By", u"Approved By", u"Issue Date", u"Date/Time"]
        for c, h in enumerate(ph, 1):
            ws3.cell(row=1, column=c, value=h)
        sty_hdr(ws3, 1, len(ph))
        for i, p in enumerate(result.get('sheet_personnel_metadata', []), 2):
            ws3.cell(row=i, column=1, value=p.get('sheet_number', ''))
            ws3.cell(row=i, column=2, value=p.get('sheet_name', ''))
            ws3.cell(row=i, column=3, value=p.get('drawn_by', ''))
            ws3.cell(row=i, column=4, value=p.get('designed_by', ''))
            ws3.cell(row=i, column=5, value=p.get('checked_by', ''))
            ws3.cell(row=i, column=6, value=p.get('approved_by', ''))
            ws3.cell(row=i, column=7, value=p.get('sheet_issue_date', ''))
            ws3.cell(row=i, column=8, value=p.get('date_time_stamp', ''))
        auto_w(ws3, len(ph))
        ws3.freeze_panes = 'A2'
        ws3.auto_filter.ref = ws3.dimensions

        # === SHEET 4: Action List ===
        ws4 = wb.create_sheet("Action List")
        ah = [u"Öncelik", u"Kategori", u"Sorun", u"Aksiyon", u"Kritik?", u"Tip"]
        for c, h in enumerate(ah, 1):
            ws4.cell(row=1, column=c, value=h)
        sty_hdr(ws4, 1, len(ah))
        for i, act in enumerate(result.get('action_list', []), 2):
            ws4.cell(row=i, column=1, value=act.get('priority', ''))
            ws4.cell(row=i, column=2, value=act.get('category', ''))
            ws4.cell(row=i, column=3, value=act.get('issue', ''))
            ws4.cell(row=i, column=4, value=act.get('action', ''))
            ws4.cell(row=i, column=5, value=u"Evet" if act.get('is_critical') else u"Hayır")
            ws4.cell(row=i, column=6, value=act.get('check_type', ''))
            if act.get('is_critical'):
                for col in range(1, len(ah) + 1):
                    ws4.cell(row=i, column=col).fill = red
        auto_w(ws4, len(ah))
        ws4.freeze_panes = 'A2'
        ws4.auto_filter.ref = ws4.dimensions

        # === SHEET 5: Raw Metrics ===
        ws5 = wb.create_sheet("Raw Metrics")
        ws5.cell(row=1, column=1, value=u"Metrik")
        ws5.cell(row=1, column=2, value=u"Değer")
        sty_hdr(ws5, 1, 2)
        r = 2
        for key in sorted(result.get('metrics', {}).keys()):
            ws5.cell(row=r, column=1, value=key)
            ws5.cell(row=r, column=2, value=utext(result['metrics'][key]))
            r += 1
        auto_w(ws5, 2)

        # === SHEET 6: QC Scoring Breakdown ===
        ws6 = wb.create_sheet("QC Scoring")
        gsd = result.get('general_score_data', {})
        bfs = result.get('bim_form_score', {})
        scoring_rows = [
            (u"GENEL SKOR", u"{0}/100".format(result.get('total_score', 0))),
            (u"", u""),
            (u"BIM Formu Skoru", u"{0}/100".format(gsd.get('bim_form_score', 0))),
            (u"BIM Formu Ağırlık", u"%{0}".format(int(gsd.get('bim_weight', 0.5) * 100))),
            (u"BIM Katkı", u"{0}".format(gsd.get('bim_component', 0))),
            (u"", u""),
            (u"Diğer QC Skoru", u"{0}/100".format(gsd.get('other_qc_score', 0))),
            (u"Diğer QC Ağırlık", u"%{0}".format(int(gsd.get('qc_weight', 0.5) * 100))),
            (u"QC Katkı", u"{0}".format(gsd.get('qc_component', 0))),
            (u"", u""),
            (u"Toplam BIM Madde", bfs.get('total_items', 0)),
            (u"Manuel Teyit Bekleyen", bfs.get('pending_count', 0)),
            (u"Değerlendirilen Madde", bfs.get('total_items', 0) - bfs.get('pending_count', 0)),
            (u"", u""),
            (u"NOT", u"Title block verileri kalite skoruna dahil değildir"),
        ]
        # AUTO / SEMI / MANUAL breakdown
        bfr = result.get('bim_form_results', [])
        auto_items = [r for r in bfr if r.get('check_type') == 'AUTO']
        semi_items = [r for r in bfr if r.get('check_type') == 'SEMI_AUTO']
        manual_items = [r for r in bfr if r.get('check_type') == 'MANUAL']
        auto_scored = [r for r in auto_items if r.get('score', -1) >= 0]
        semi_scored = [r for r in semi_items if r.get('score', -1) >= 0]
        manual_scored = [r for r in manual_items if r.get('score', -1) >= 0]

        def _avg(items):
            if not items: return 0
            return round(sum(r['score'] for r in items) / float(len(items)), 2)

        scoring_rows += [
            (u"", u""),
            (u"AUTO Madde Sayısı", len(auto_items)),
            (u"AUTO Ortalama Puan", u"{0}/5".format(_avg(auto_scored))),
            (u"SEMI_AUTO Madde Sayısı", len(semi_items)),
            (u"SEMI_AUTO Değerlendirilen", len(semi_scored)),
            (u"MANUAL Madde Sayısı", len(manual_items)),
            (u"MANUAL Onaylanan", len(manual_scored)),
        ]
        for i, (l, v) in enumerate(scoring_rows, 1):
            ws6.cell(row=i, column=1, value=l).font = Font(bold=True) if l else Font()
            ws6.cell(row=i, column=2, value=utext(v))
        auto_w(ws6, 2)

        wb.save(xlsx_path)
        return True, xlsx_path
    except Exception as exc:
        msg = u"Excel hatası: {0}".format(exc)
        logger.error(msg)
        return False, msg
